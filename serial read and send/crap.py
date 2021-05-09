from openpyxl import *
from tkinter import *
from tkinter import ttk


def menu_1():
    print("1.player menu opened")
    print("2.new Team Sign-up")
    # opening the Team excel file
    wb = load_workbook('PlayerReg.xlsx')
    print("2.Loading PlayerExcal")
    # create the sheet object
    sheet1 = wb.active


    def excel():

        # resize the width of columns in
        # excel spreadsheet
        sheet1.column_dimensions['A'].width = 25
        sheet1.column_dimensions['B'].width = 25
        sheet1.column_dimensions['C'].width = 25
        sheet1.column_dimensions['D'].width = 25
        sheet1.column_dimensions['E'].width = 10


        # write given data to an excel spreadsheet
        # at particular location
        sheet1.cell(row=1, column=1).value = "Full Name"
        sheet1.cell(row=1, column=2).value = "Team Name"
        sheet1.cell(row=1, column=3).value = "Phone Number"
        sheet1.cell(row=1, column=4).value = "Email"
        sheet1.cell(row=1, column=5).value = "Age"

    # Function to set focus (cursor)
    def focus1(event):
        # set focus on the course_field box
        team_field.focus_set()


    # Function to set focus
    def focus2(event):
        # set focus on the phone_field box
        phone_field.focus_set()


    # Function to set focus
    def focus3(event):
        # set focus on the email_field box
        email_field.focus_set()


    # Function to set focus
    def focus4(event):
        # set focus on the contact_no_field box
        age_field.focus_set()



    # Function for clearing the
    # contents of text entry boxes
    def clear():

        # clear the content of text entry box
        name_field.delete(0, END)
        team_field.delete(0, END)
        phone_field.delete(0, END)
        email_field.delete(0, END)
        age_field.delete(0, END)



    # Function to take data from GUI
    # window and write to an excel file
    def insert():

        # if user not fill any entry
        # then print "empty input"
        if (name_field.get() == "" and
            team_field.get() == "" and
            phone_field.get() == "" and
            email_field.get() == "" and
            age_field.get() == ""):
            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet1.max_row
            current_column = sheet1.max_column

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet1.cell(row=current_row + 1, column=1).value = name_field.get()
            sheet1.cell(row=current_row + 1, column=2).value = team_field.get()
            sheet1.cell(row=current_row + 1, column=3).value = phone_field.get()
            sheet1.cell(row=current_row + 1, column=4).value = email_field.get()
            sheet1.cell(row=current_row + 1, column=5).value = age_field.get()


            # save the file
            wb.save('PlayerReg.xlsx')

            # set focus on the name_field box
            name_field.focus_set()

            # call the clear() function
            clear()


    # Driver code
    if __name__ == "__main__":

        # create a GUI window
        root = Tk()

        # set the background colour of GUI window
        root.configure(background='white')

        # set the title of GUI window
        root.title("Player Registration")

        # set the configuration of GUI window
        root.geometry("500x200")

        excel()

        # create a Form label
        heading = Label(tab1, text="New Player", bg="white")

        # create a Name label
        name = Label(tab1, text="Full Name", bg="white")

        # create a Team label
        team = Label(tab1, text="Team Name", bg="white")

        # create a phone label
        phone = Label(tab1, text="Contact Number", bg="white")

        # create a Form No. label
        email = Label(tab1, text="Email", bg="white")

        # create a age label
        age = Label(tab1, text="Age", bg="white")



        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        team.grid(row=2, column=0)
        phone.grid(row=3, column=0)
        email.grid(row=4, column=0)
        age.grid(row=5, column=0)


        # create a text entry box
        # for typing the information
        name_field = Entry(root)
        team_field = Entry(root)
        phone_field = Entry(root)
        email_field = Entry(root)
        age_field = Entry(root)


        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        name_field.bind("<Return>", focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        team_field.bind("<Return>", focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        phone_field.bind("<Return>", focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        email_field.bind("<Return>", focus4)

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        name_field.grid(row=1, column=1, ipadx="100")
        team_field.grid(row=2, column=1, ipadx="100")
        phone_field.grid(row=3, column=1, ipadx="100")
        email_field.grid(row=4, column=1, ipadx="100")
        age_field.grid(row=5, column=1, ipadx="100")


        # call excel function
        excel()

        # create a Submit Button and place into the root window

        ttk.Button(tab1, text="Submit", command=insert).grid(column=1, row=8)
        print("2.Saved new player")

        # start the GUI
        root.mainloop()

def menu_2():
    print("1.team menu opened")
    print("3.New Team Sign-up")
    # opening the existing excel file
    wb = load_workbook('TeamReg.xlsx')
    print("3.loading TeamExcal")

    # create the sheet object
    sheet2 = wb.active


    def excel():

        # resize the width of columns in
        # excel spreadsheet
        sheet2.column_dimensions['A'].width = 25
        sheet2.column_dimensions['B'].width = 25
        sheet2.column_dimensions['C'].width = 25
        sheet2.column_dimensions['D'].width = 25
        sheet2.column_dimensions['E'].width = 10


        # write given data to an excel spreadsheet
        # at particular location
        sheet2.cell(row=1, column=1).value = "Team Name"
        sheet2.cell(row=1, column=2).value = "Number of players"
        sheet2.cell(row=1, column=3).value = "Manager Name"
        sheet2.cell(row=1, column=4).value = "Manager Contact"
        sheet2.cell(row=1, column=5).value = "Age Group"

    # Function to set focus (cursor)
    def focus1(event):
        # set focus on the course_field box
        number_field.focus_set()


    # Function to set focus
    def focus2(event):
        # set focus on the phone_field box
        manager_field.focus_set()


    # Function to set focus
    def focus3(event):
        # set focus on the email_field box
        contact_field.focus_set()


    # Function to set focus
    def focus4(event):
        # set focus on the contact_no_field box
        ageG_field.focus_set()



    # Function for clearing the
    # contents of text entry boxes
    def clear():

        # clear the content of text entry box
        team1_field.delete(0, END)
        number_field.delete(0, END)
        manager_field.delete(0, END)
        contact_field.delete(0, END)
        ageG_field.delete(0, END)



    # Function to take data from GUI
    # window and write to an excel file
    def insert():

        # if user not fill any entry
        # then print "empty input"
        if (team1_field.get() == "" and
            number_field.get() == "" and
            manager_field.get() == "" and
            contact_field.get() == "" and
            ageG_field.get()):
            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet2.max_row
            current_column = sheet2.max_column

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet2.cell(row=current_row + 1, column=1).value = team1_field.get()
            sheet2.cell(row=current_row + 1, column=2).value = number_field.get()
            sheet2.cell(row=current_row + 1, column=3).value = manager_field.get()
            sheet2.cell(row=current_row + 1, column=4).value = contact_field.get()
            sheet2.cell(row=current_row + 1, column=5).value = ageG_field.get()


            # save the file
            wb.save('TeamReg.xlsx')

            # set focus on the name_field box
            team1_field.focus_set()

            # call the clear() function
            clear()


    # Driver code
    if __name__ == "__main__":

        # create a GUI window
        root = Tk()

        # set the background colour of GUI window
        root.configure(background='white')

        # set the title of GUI window
        root.title("Player Registration")

        # set the configuration of GUI window
        root.geometry("500x200")

        excel()

        # create a Form label
        team1 = Label(tab1, text="Team Name", bg="white")

        # create a Name label
        number = Label(tab1, text="Number of players", bg="white")

        # create a Team label
        manager = Label(tab1, text="Manager Name", bg="white")

        # create a phone label
        contact = Label(tab1, text="Manager Contact", bg="white")

        # create a Form No. label
        ageG = Label(tab1, text="Age Group", bg="white")


        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        team1.grid(row=1, column=0)
        number.grid(row=2, column=0)
        manager.grid(row=3, column=0)
        contact.grid(row=4, column=0)
        ageG.grid(row=5, column=0)


        # create a text entry box
        # for typing the information
        team1_field = Entry(root)
        number_field = Entry(root)
        manager_field = Entry(root)
        contact_field = Entry(root)
        ageG_field = Entry(root)


        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        team1_field.bind("<Return>", focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        number_field.bind("<Return>", focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        manager_field.bind("<Return>", focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        contact_field.bind("<Return>", focus4)



        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        team1_field.grid(row=1, column=1, ipadx="100")
        number_field.grid(row=2, column=1, ipadx="100")
        manager_field.grid(row=3, column=1, ipadx="100")
        contact_field.grid(row=4, column=1, ipadx="100")
        ageG_field.grid(row=5, column=1, ipadx="100")


        # call excel function
        excel()

        # create a Submit Button and place into the root window

        ttk.Button(tab1, text="Submit", command=insert).grid(column=1, row=8)
        print("3.Saved new team")
        # start the GUI
        root.mainloop()



def ViewPlayers():
    print("not working")


def menu():
    #Window
    BendigoFutsal = Tk()
    BendigoFutsal.title("Bendigo Futsal")
    BendigoFutsal.configure(background='white')
    BendigoFutsal.geometry("300x200")
    print("1.opened Menu")

    # Labels
    Label_1 = Label(text="Bendigo Futsal")
    Label_1.grid(row=1, column=2)
    Label_1.configure(bg='white', fg='Black', font="Verdana 35 bold")

    Label_2 = Label(text="Coded by Harry Pitson")
    Label_2.grid(row=9, column=2)
    Label_2.configure(bg='white', fg='Gray', font="Verdana 10 italic")

    # Buttons
    ttk.Button(BendigoFutsal, text="Player Registration", command=menu_1, width=20).grid(column=2, row=2)
    ttk.Button(BendigoFutsal, text="Team Registration", command=menu_2, width=20).grid(column=2, row=4)
    ttk.Button(BendigoFutsal, text="View Players", command=ViewPlayers, width=20).grid(column=2, row=5)
    ttk.Button(BendigoFutsal, text="View Teams", command=ViewPlayers, width=20).grid(column=2, row=6)

    def Close():
        #Ending the loop For the menu.
        #Therfor closing the window.
        print("1.Closed Menu")
        BendigoFutsal.destroy()

    ttk.Button(BendigoFutsal, text="Exit", command=Close).grid(column=2, row=10)


    BendigoFutsal.mainloop()
menu()